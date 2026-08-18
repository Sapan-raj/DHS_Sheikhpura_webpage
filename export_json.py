#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reads Sheikhpura_Health_PIP_Website_Database.xlsx and writes
assets/data/portal-data.json

This JSON is the OFFLINE FALLBACK data source. It is generated from the
workbook — never hand-edited. It has byte-for-byte the same shape that the
Google Apps Script Web App returns, so the frontend has exactly one contract
regardless of which source is active.

Run:  python export_json.py
"""
import json, os, datetime
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "Sheikhpura_Health_PIP_Website_Database.xlsx")
OUT = os.path.join(ROOT, "assets", "data", "portal-data.json")

# sheet name -> json key. Must match apps-script/Code.gs SHEET_MAP exactly.
SHEET_MAP = {
    "Settings":            "settings",
    "Navigation":          "navigation",
    "Financial_Years":     "financialYears",
    "Program_Categories":  "categories",
    "Programs_FMR":        "programs",
    "Documents":           "documents",
    "Home_Content":        "home",
    "Important_Links":     "links",
    "Notices":             "notices",
    "Contact_Information": "contacts",
    "Footer":              "footer",
    "Post_Categories":     "postCategories",
    "Posts":               "posts",
    "Post_Media":          "postMedia",
    "Program_Benefits":    "benefits",
}


def cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def main():
    wb = load_workbook(XLSX, data_only=True)
    out = {
        "meta": {
            "generated": datetime.datetime.now().replace(microsecond=0).isoformat(),
            "source": "xlsx-export",
            "version": datetime.datetime.now().strftime("%Y%m%d%H%M%S"),
        }
    }

    for sname, key in SHEET_MAP.items():
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() for h in rows[0] if h is not None]
        recs = []
        for r in rows[1:]:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            recs.append({h: cell(r[i]) for i, h in enumerate(headers)})
        out[key] = recs

    # Settings collapse to a flat map — the shape the frontend actually wants.
    out["settings"] = {s["Setting_Key"]: s["Setting_Value"] for s in out["settings"]}

    # This file is served publicly, so it must never carry unpublished content.
    # Same rule as publicView() in apps-script/Code.gs: Drafts never leave, and a
    # Scheduled post only appears once its moment has passed.
    now = datetime.datetime.now()

    def _when(v):
        s = str(v or "").strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(s[:len(fmt) + 2].strip(), fmt)
            except ValueError:
                continue
        return None

    def _visible(p):
        st = str(p.get("Status", "")).strip().lower()
        if st in ("published", "archived"):
            return True
        if st == "scheduled":
            w = _when(p.get("Scheduled_Date"))
            return bool(w and w <= now)
        return False

    held = [p for p in out["posts"] if not _visible(p)]
    out["posts"] = [p for p in out["posts"] if _visible(p)]
    live_ids = {p["Post_ID"] for p in out["posts"]}
    out["postMedia"] = [m for m in out["postMedia"] if m.get("Post_ID") in live_ids]
    out["_withheld"] = len(held)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    size = os.path.getsize(OUT) / 1024
    print(f"Wrote {OUT}  ({size:.1f} KB)")
    for k in SHEET_MAP.values():
        n = len(out[k]) if isinstance(out[k], list) else len(out[k])
        print(f"   {k:<16} {n:>4}")

    # quick contract check
    fy = {f["Year_ID"] for f in out["financialYears"]}
    bad = [p["Program_ID"] for p in out["programs"] if p["Year_ID"] not in fy]
    if held:
        print(f"\nWithheld from the public snapshot ({len(held)}):")
        for p in held:
            print(f"   {p['Post_ID']:<9} {p['Status']:<10} {str(p['Title'])[:52]}")

    print(f"\nFK check programs->financialYears: {'PASS' if not bad else 'FAIL ' + str(bad[:5])}")
    per = {}
    for p in out["programs"]:
        per[p["Year_ID"]] = per.get(p["Year_ID"], 0) + 1
    print("FMR heads per year:", per)


if __name__ == "__main__":
    main()
