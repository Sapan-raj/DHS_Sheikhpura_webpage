#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generates Sheikhpura_Health_PIP_Website_Database.xlsx

Google-Sheets-ready master database for the Sheikhpura District Health PIP Portal.
Rules enforced:
  - one record per row, no merged cells, no formulas
  - every sheet has a single header row (frozen)
  - unique IDs, relational integrity between sheets
  - URLs in dedicated columns, metadata separate from URLs
  - controlled vocabularies kept in _Lists
Run:  python build_database.py
"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from benefits_data import BENEFITS, BENEFIT_COLUMNS, BENEFIT_TYPES

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Sheikhpura_Health_PIP_Website_Database.xlsx")

NAVY   = "1B3A5C"
ORANGE = "F2994A"
GREY   = "F2F4F7"
MARK   = "NEEDS MANUAL INPUT"

hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
hdr_fill = PatternFill("solid", fgColor=NAVY)
thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def sheet(wb, name, headers, rows, widths=None, note=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, border
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for r in rows:
        ws.append(list(r))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cl = ws.cell(row=r, column=c)
            cl.border = border
            cl.alignment = Alignment(vertical="top", wrap_text=False)
            cl.font = Font(size=10, name="Calibri")
            if cl.value == MARK:
                cl.font = Font(size=10, name="Calibri", bold=True, color="B42318")
    for i, h in enumerate(headers, start=1):
        w = (widths or {}).get(h, max(12, min(42, len(str(h)) + 6)))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row,1)}"
    if note:
        ws.sheet_properties.tabColor = ORANGE
    return ws


# ───────────────────────── 1. SETTINGS ─────────────────────────
SETTINGS = [
    # Key, Label, Value, Group, Notes
    ("district_name",        "District Name",              "Sheikhpura", "Identity", "Appears in header and page titles"),
    ("district_name_hi",     "District Name (Hindi)",      "शेखपुरा", "Identity", ""),
    ("district_code",        "District Code (LGD)",        "225", "Identity", "Bihar LGD district code"),
    ("state_name",           "State",                      "Bihar", "Identity", ""),
    ("site_title",           "Website Title",              "District Health Society, Sheikhpura", "Identity", "Browser tab + header line 1"),
    ("site_title_hi",        "Website Title (Hindi)",      "जिला स्वास्थ्य समिति, शेखपुरा", "Identity", ""),
    ("site_subtitle",        "Subtitle",                   "Health Services for the People of Sheikhpura", "Identity", "Header line 2"),
    ("site_subtitle_hi",     "Subtitle (Hindi)",           "शेखपुरा की जनता के लिए स्वास्थ्य सेवाएं", "Identity", ""),
    ("parent_body",          "Parent Body",                "State Health Society, Bihar  |  National Health Mission", "Identity", ""),
    ("logo_url",             "Logo URL",                   "assets/img/emblem.svg", "Branding", "Leave as-is to use the built-in emblem"),
    ("favicon_url",          "Favicon URL",                "assets/img/favicon.svg", "Branding", ""),
    ("primary_color",        "Primary Colour",             "#1B3A5C", "Branding", "Hex. Header, nav, footer"),
    ("accent_color",         "Accent Colour",              "#F2994A", "Branding", "Hex. Hover, highlights"),
    ("current_financial_year", "Current Financial Year",   "2026-27", "Content", "Must match a Financial_Year value in Financial_Years"),
    ("site_status",          "Website Status",             "Live", "Content", "Live | Maintenance. Maintenance shows a banner"),
    ("maintenance_message",  "Maintenance Message",        "The portal is undergoing scheduled maintenance. Data shown may not be current.", "Content", ""),
    ("contact_email",        "Primary Email",              MARK, "Contact", "e.g. dhs-sheikhpura-bih@gov.in — confirm official ID"),
    ("contact_phone",        "Office Phone",               MARK, "Contact", "STD code + number"),
    ("helpline_104",         "Health Helpline",            "104", "Contact", "Toll free"),
    ("helpline_102",         "Ambulance Helpline",         "102", "Contact", "Toll free"),
    ("office_address",       "Office Address",             "Office of the Civil Surgeon, Sadar Hospital Campus, Sheikhpura, Bihar - 811105", "Contact", "Verify PIN before go-live"),
    ("footer_copyright",     "Copyright Line",             "© 2026 District Health Society, Sheikhpura, Government of Bihar. All rights reserved.", "Footer", ""),
    ("footer_credit",        "Developed By Line",          "Designed & Developed by District IT Cell, Sheikhpura", "Footer", ""),
    ("footer_about",         "Footer About Text",          "The District Health Society, Sheikhpura implements the National Health Mission across the district. This portal publishes the health services available free to every resident, along with the events, notices and contacts you may need.", "Footer", ""),
    ("social_facebook",      "Facebook URL",               "", "Social", "Blank hides the icon"),
    ("social_twitter",       "Twitter / X URL",            "", "Social", "Blank hides the icon"),
    ("social_youtube",       "YouTube URL",                "", "Social", "Blank hides the icon"),
    ("google_maps_url",      "Google Maps URL",            "https://www.google.com/maps/search/Sadar+Hospital+Sheikhpura+Bihar", "Contact", ""),
    ("show_visitor_counter", "Show Visitor Counter",       "Yes", "Footer", "Yes | No"),
    ("records_per_page",     "Records Per Page",           "25", "Behaviour", "Pagination size on Documents page"),
    ("cache_minutes",        "Browser Cache (minutes)",    "30", "Behaviour", "How long the site reuses data before re-fetching"),
    ("enable_hindi",         "Enable Hindi Toggle",        "Yes", "Behaviour", "Yes | No"),
]

# ───────────────────────── 2. NAVIGATION ─────────────────────────
NAV = [
    ("NAV01", "Home",              "होम",                    "index.html",      "", 1, "Yes", "Internal", "home",     "_self"),
    ("NAV02", "Free Services",     "नि:शुल्क सेवाएं",           "benefits.html",   "", 2, "Yes", "Internal", "heart",    "_self"),
    ("NAV03", "Health Programmes", "स्वास्थ्य कार्यक्रम",        "programmes.html", "", 3, "Yes", "Internal", "hospital", "_self"),
    ("NAV04", "Events & News",     "कार्यक्रम एवं समाचार",      "events.html",     "", 4, "Yes", "Internal", "bell",     "_self"),
    ("NAV05", "Notices",           "सूचनाएँ",                  "notices.html",    "", 5, "Yes", "Internal", "bell",     "_self"),
    ("NAV06", "Documents",         "दस्तावेज़",                "documents.html",  "", 6, "Yes", "Internal", "download", "_self"),
    ("NAV07", "Contact Us",        "संपर्क करें",              "contact.html",    "", 7, "Yes", "Internal", "phone",    "_self"),
    ("NAV08", "BHAVYA Portal",     "भव्य पोर्टल",              "https://mera.bhavyabiharhealth.in/", "", 8, "Yes", "External", "external", "_blank"),
]

# ───────────────────────── 3. FINANCIAL YEARS ─────────────────────────
FY = [
    ("FY2627", "2026-27", "F.Y. 2026-2027", 2026, 2027, "Yes", "Active",   1),
    ("FY2526", "2025-26", "F.Y. 2025-2026", 2025, 2026, "No",  "Active",   2),
    ("FY2425", "2024-25", "F.Y. 2024-2025", 2024, 2025, "No",  "Active",   3),
    ("FY2324", "2023-24", "F.Y. 2023-2024", 2023, 2024, "No",  "Archived", 4),
]

# ───────────────────────── 5. PROGRAM CATEGORIES (global master) ─────────────────────────
CATS = [
    ("CAT01", "RCH Flexible Pool (including RI, IPPI, NIDDCP)", "RCH",    "Reproductive, Maternal, Newborn, Child and Adolescent Health, Routine Immunisation, Pulse Polio and iodine deficiency control.", "heart",  1, "Active"),
    ("CAT02", "NDCP Flexi Pool",                                "NDCP",   "National Disease Control Programmes — surveillance, vector-borne disease, leprosy, tuberculosis, viral hepatitis and rabies.", "shield", 2, "Active"),
    ("CAT03", "NCD Flexi Pool",                                 "NCD",    "Non-Communicable Disease programmes — blindness, mental health, elderly care, tobacco control, diabetes and cardiovascular disease.", "pulse", 3, "Active"),
    ("CAT04", "Health System Strengthening (HSS) - Urban",      "HSS(U)", "Urban health system strengthening under NUHM — urban primary healthcare, community engagement, quality assurance and human resources.", "city", 4, "Active"),
    ("CAT05", "Health System Strengthening (HSS) - Rural",      "HSS",    "Rural health system strengthening — comprehensive primary healthcare, blood services, referral transport, infrastructure, HR and IT systems.", "hospital", 5, "Active"),
]

# ───────────────── 6. PROGRAMS / FMR — year-scoped (the critical table) ─────────────────
FMR_2526 = [
    ("CAT01", "RCH.1",     "Maternal Health", "मातृ स्वास्थ्य"),
    ("CAT01", "RCH.2",     "PC-PNDT", "पीसी-पीएनडीटी"),
    ("CAT01", "RCH.3",     "Child Health", "बाल स्वास्थ्य"),
    ("CAT01", "RCH.4",     "Immunization", "टीकाकरण"),
    ("CAT01", "RCH.5",     "Adolescent Health", "किशोर स्वास्थ्य"),
    ("CAT01", "RCH.6",     "Family Planning", "परिवार नियोजन"),
    ("CAT01", "RCH.7",     "Nutrition", "पोषण"),
    ("CAT01", "RCH.8",     "National Iodine Deficiency Disorders Control Programme (NIDDCP)", ""),
    ("CAT02", "NDCP.1",    "Integrated Disease Surveillance Programme (IDSP)", ""),
    ("CAT02", "NDCP.2",    "National Vector Borne Disease Control Programme (NVBDCP)", ""),
    ("CAT02", "NDCP.3",    "National Leprosy Eradication Programme (NLEP)", ""),
    ("CAT02", "NDCP.4",    "National Tuberculosis Elimination Programme (NTEP)", ""),
    ("CAT02", "NDCP.5",    "National Viral Hepatitis Control Programme (NVHCP)", ""),
    ("CAT02", "NDCP.6",    "National Rabies Control Programme (NRCP)", ""),
    ("CAT02", "NDCP.8",    "State specific Initiatives and Innovations", ""),
    ("CAT03", "NCD.1",     "National Program for Control of Blindness and Vision Impairment (NPCB+VI)", ""),
    ("CAT03", "NCD.2",     "National Mental Health Program (NMHP)", ""),
    ("CAT03", "NCD.3",     "National Programme for Health Care for the Elderly (NPHCE)", ""),
    ("CAT03", "NCD.4",     "National Tobacco Control Programme (NTCP)", ""),
    ("CAT03", "NCD.5",     "National Programme for Prevention and Control of Diabetes, Cardiovascular Disease and Stroke (NPCDCS)", ""),
    ("CAT03", "NCD.6",     "Pradhan Mantri National Dialysis Programme (PMNDP)", ""),
    ("CAT03", "NCD.7",     "National Program for Climate Change and Human Health (NPCCHH)", ""),
    ("CAT03", "NCD.8",     "National Oral Health Programme (NOHP)", ""),
    ("CAT03", "NCD.10",    "National Programme for Prevention and Control of Fluorosis (NPPCF)", ""),
    ("CAT03", "NCD.11",    "National Programme for Prevention and Control of Deafness (NPPCD)", ""),
    ("CAT03", "NCD.13",    "State specific Programme Interventions", ""),
    ("CAT04", "HSS(U).1",  "Comprehensive Primary Healthcare (CPHC)", ""),
    ("CAT04", "HSS(U).2",  "Community Engagement", ""),
    ("CAT04", "HSS(U).3",  "Public Health Institutions as per IPHS norms", ""),
    ("CAT04", "HSS(U).4",  "Quality Assurance", ""),
    ("CAT04", "HSS(U).5",  "HRH", ""),
    ("CAT04", "HSS(U).6",  "Technical Assistance", ""),
    ("CAT04", "HSS(U).7",  "Access", ""),
    ("CAT04", "HSS(U).8",  "Innovation", ""),
    ("CAT04", "HSS(U).9",  "Untied Grants", ""),
    ("CAT05", "HSS.1",     "Comprehensive Primary Healthcare (CPHC)", ""),
    ("CAT05", "HSS.2",     "Blood Services & Disorders", ""),
    ("CAT05", "HSS.3",     "Community Engagement", ""),
    ("CAT05", "HSS.4",     "Public Health Institutions as per IPHS norms", ""),
    ("CAT05", "HSS.5",     "Referral Transport", ""),
    ("CAT05", "HSS.6",     "Quality Assurance", ""),
    ("CAT05", "HSS.7",     "Other Initiatives to improve access", ""),
    ("CAT05", "HSS.8",     "Inventory management", ""),
    ("CAT05", "HSS.9",     "HRH", ""),
    ("CAT05", "HSS.10",    "Enhancing HR", ""),
    ("CAT05", "HSS.11",    "Technical Assistance", ""),
    ("CAT05", "HSS.12",    "IT interventions and systems", ""),
    ("CAT05", "HSS.13",    "Innovation", ""),
    ("CAT05", "HSS.14",    "Untied Grants", ""),
]

# FY 2026-27 differs — this is the whole reason FMR rows are year-scoped
FMR_2627 = [r for r in FMR_2526 if r[1] not in ("NDCP.8",)]
FMR_2627 += [
    ("CAT02", "NDCP.7",  "National Programme on Containment of Anti-Microbial Resistance (AMR)", ""),
    ("CAT02", "NDCP.9",  "State specific Initiatives and Innovations", ""),
    ("CAT03", "NCD.9",   "National Programme for Palliative Care (NPPC)", ""),
    ("CAT03", "NCD.12",  "National Programme for Prevention and Control of Burn Injuries (NPPCBI)", ""),
    ("CAT05", "HSS.15",  "Infrastructure Maintenance", ""),
    ("CAT05", "HSS.16",  "Programme Management", ""),
    ("CAT05", "HSS.17",  "Ayushman Arogya Mandir - Operational Cost", ""),
    ("CAT05", "HSS.18",  "Biomedical Equipment Management & Maintenance", ""),
    ("CAT05", "HSS.19",  "Digital Health Initiatives", ""),
    ("CAT05", "HSS.20",  "Grievance Redressal Systems", ""),
    ("CAT05", "HSS.21",  "State specific Health System Interventions", ""),
]

CAT_ORDER = {"CAT01": 1, "CAT02": 2, "CAT03": 3, "CAT04": 4, "CAT05": 5}


def fmr_sort_key(row):
    cat, code = row[0], row[1]
    num = int(code.split(".")[-1])
    return (CAT_ORDER[cat], num)


NODAL = {
    "CAT01": "District Programme Manager / DCM",
    "CAT02": "District Vector Borne Disease Control Officer",
    "CAT03": "District NCD Nodal Officer",
    "CAT04": "Urban Health Nodal Officer",
    "CAT05": "District Programme Manager",
}


def build_programs():
    rows, n = [], 0
    for yid, data in (("FY2627", sorted(FMR_2627, key=fmr_sort_key)),
                      ("FY2526", sorted(FMR_2526, key=fmr_sort_key)),
                      ("FY2425", sorted(FMR_2526, key=fmr_sort_key))):
        for order, (cat, code, name, name_hi) in enumerate(data, start=1):
            n += 1
            rows.append((
                f"PRG{n:04d}", yid, cat, code, name, name_hi,
                f"{name} services provided across the six blocks of Sheikhpura district "
                f"through government health facilities.",
                NODAL[cat], order, "Active",
            ))
    return rows


PROGRAMS = build_programs()

# ───────────────── 4. PIP_DOCUMENTS — year-level headline strip ─────────────────
# ───────── 7. DOCUMENTS — category allocations, guidelines, programme & general files ─────────
def build_documents():
    """Citizen-facing documents only. Budget allocation and guideline files were
    removed at the District Magistrate's direction — this portal publishes what
    the public can use, not the district's financial plan."""
    rows = []
    extra = [
        ("FY2627", "CAT01", "PRG0001", "Maternal Health Operational Guideline 2026-27", "Programme Guideline",
         "Operational guideline for maternal health services including JSY and PMSMA.", "PDF", "Yes"),
        ("FY2627", "CAT02", "",        "NIKSHAY Reporting Format",                      "Format",
         "Monthly reporting format for TB notification and Ni-kshay Poshan Yojana.",     "XLSX", "No"),
        ("FY2627", "CAT05", "",        "Ayushman Arogya Mandir Monthly Report Format",   "Format",
         "Standard monthly reporting format for AAM/HWC service delivery.",              "XLSX", "Yes"),
        ("FY2627", "",      "",        "District Health Society Meeting Circular",       "Circular",
         "Circular convening the quarterly District Health Society governing body meeting.", "PDF", "No"),
        ("FY2526", "CAT03", "",        "NCD Screening Progress Report 2025-26",          "Report",
         "Block-wise population based screening progress for F.Y. 2025-26.",             "PDF", "No"),
    ]
    for n, (yid, cid, pid, title, dtype, desc, ftype, feat) in enumerate(extra, start=1):
        rows.append((f"DOC{n:04d}", yid, cid, pid, title, dtype, desc, MARK, ftype, "", "", n, "Active", feat))
    return rows


DOCUMENTS = build_documents()

# ───────────────────────── 8. HOME PAGE CONTENT ─────────────────────────
HOME = [
    ("hero_title",     "hero",    "District Health Society, Sheikhpura", "जिला स्वास्थ्य समिति, शेखपुरा", "Health services for every resident", "", "", "", "", 1, "Active"),
    ("hero_subtitle",  "hero",    "", "", "", "Find the free health services you are entitled to, where to go for them, and what is happening in the district this month.", "", "benefits.html", "See free services", 2, "Active"),
    ("hero_cta2",      "hero",    "", "", "", "", "", "events.html", "What's happening", 3, "Active"),
    ("notice_banner",  "banner",  "Announcement", "घोषणा", "", "Every service listed on this portal is free at government health facilities. If anyone asks you to pay, call 104.", "", "benefits.html", "Know your rights", 4, "Active"),
    ("sec_benefits",   "section", "Free Health Services for You", "आपके लिए नि:शुल्क स्वास्थ्य सेवाएं", "Know what you are entitled to", "Every one of these is free at government health facilities in Sheikhpura. Nobody may charge you for them.", "", "benefits.html", "See all free services", 5, "Active"),
    ("stat_blocks",    "stat",    "Blocks", "प्रखंड", "", "6", "grid", "", "", 6, "Active"),
    ("stat_benefits",  "stat",    "Free Services", "नि:शुल्क सेवाएं", "", "", "heart", "benefits.html", "", 7, "Active"),
    ("stat_programmes","stat",    "Health Programmes", "स्वास्थ्य कार्यक्रम", "", "", "hospital", "programmes.html", "", 8, "Active"),
    ("stat_helpline",  "stat",    "Health Helpline", "स्वास्थ्य हेल्पलाइन", "", "104", "phone", "contact.html", "", 9, "Active"),
    ("sec_whatsnew",   "section", "What's New", "नया क्या है", "Events, campaigns and announcements", "Health campaigns, camps, training programmes and achievements from across Sheikhpura district.", "", "events.html", "View all", 10, "Active"),
    ("sec_programmes", "section", "Health Programmes", "स्वास्थ्य कार्यक्रम", "What the district runs for you", "The health programmes running in Sheikhpura and the services each one provides free of cost.", "", "programmes.html", "View all programmes", 11, "Active"),
    ("sec_notices",    "section", "Notices & Announcements", "सूचनाएँ एवं घोषणाएँ", "From the District Health Society", "", "", "notices.html", "All notices", 12, "Active"),
    ("sec_quicklinks", "section", "Quick Links", "त्वरित लिंक", "Frequently used portals", "", "", "", "", 13, "Active"),
    ("about_text",     "richtext","About this portal", "इस पोर्टल के बारे में", "", "The District Health Society, Sheikhpura implements the National Health Mission across the district. This portal exists so that every resident can find out what health services they are entitled to, where to go for them, what to carry, and who to call — without having to ask anyone. Every service listed here is provided free of cost at government health facilities.", "", "", "", 14, "Active"),
]

# ───────────────────────── 9. IMPORTANT LINKS ─────────────────────────
LINKS = [
    ("LNK01", "State Health Society, Bihar",        "https://shs.bihar.gov.in/",              "State Health Society — parent body for district health services",  "external", "State",    1,  "Active", "Yes", "Yes", "Yes"),
    ("LNK02", "National Health Mission, GoI",       "https://nhm.gov.in/",                    "National Health Mission portal",              "external", "National", 2,  "Active", "Yes", "Yes", "Yes"),
    ("LNK03", "MoHFW, Government of India",         "https://mohfw.gov.in/",                  "Ministry of Health and Family Welfare",       "external", "National", 3,  "Active", "Yes", "Yes", "No"),
    ("LNK04", "Government of Bihar",                "https://state.bihar.gov.in/",            "State government portal",                     "external", "State",    4,  "Active", "Yes", "Yes", "No"),
    ("LNK05", "Health Department, Bihar",           "https://state.bihar.gov.in/health/",     "Department of Health, Government of Bihar",   "external", "State",    5,  "Active", "Yes", "Yes", "No"),
    ("LNK06", "Sheikhpura District Portal",         "https://sheikhpura.nic.in/",             "District administration website",             "external", "District", 6,  "Active", "Yes", "Yes", "Yes"),
    ("LNK07", "BHAVYA Citizen Health Portal",       "https://mera.bhavyabiharhealth.in/",     "Bihar health citizen services platform",      "external", "Portal",   7,  "Active", "Yes", "Yes", "Yes"),
    ("LNK08", "HMIS — GoI Web Portal",              "https://hmis.mohfw.gov.in/",             "Health Management Information System",        "external", "Portal",   8,  "Active", "Yes", "Yes", "Yes"),
    ("LNK09", "RCH Portal",                         "https://rch.nhm.gov.in/RCH/",            "Reproductive and Child Health portal",        "external", "Portal",   9,  "Active", "Yes", "Yes", "No"),
    ("LNK10", "Ni-kshay (TB)",                      "https://www.nikshay.in/",                "National TB elimination reporting system",    "external", "Portal",   10, "Active", "Yes", "Yes", "No"),
    ("LNK11", "ABDM — Ayushman Bharat Digital",     "https://abdm.gov.in/",                   "Digital health mission",                      "external", "Portal",   11, "Active", "Yes", "Yes", "No"),
    ("LNK12", "BMSICL",                             "https://bmsicl.gov.in/",                 "Bihar Medical Services & Infrastructure Corp.","external","State",    12, "Active", "Yes", "Yes", "No"),
    ("LNK13", "e-Procurement Portal, Bihar",        "https://eproc2.bihar.gov.in/",           "State e-procurement",                         "external", "State",    13, "Active", "Yes", "Yes", "No"),
    ("LNK14", "IHIP — Integrated Health Info Platform","https://ihip.mohfw.gov.in/",          "Disease surveillance reporting",              "external", "Portal",   14, "Active", "Yes", "Yes", "No"),
]

# ───────────────────────── 10. NOTICES ─────────────────────────
NOTICES = [
    ("NOT01", "Free health services available at every government facility",
     "Delivery, medicines, tests, ambulance, TB treatment, dialysis, cataract surgery and much more are provided free of cost at government health facilities in Sheikhpura. If anyone asks you to pay for these, call 104 and complain.",
     "2026-08-15", "General", "High", "", "benefits.html", "Yes", "Yes", "Active", 1, ""),

    ("NOT02", "Ayushman Arogya Mandir now open in every panchayat",
     "Your nearest Health Sub-Centre now provides twelve health services free — pregnancy care, child care, immunisation, treatment of common illnesses, screening for blood pressure and diabetes, and free video consultation with a doctor.",
     "2026-08-10", "General", "High", "", "benefits.html", "Yes", "Yes", "Active", 2, ""),

    ("NOT03", "Free screening for blood pressure, diabetes and cancer",
     "Everyone aged 30 and above can get free screening for high blood pressure, diabetes and common cancers. Your ASHA can test you at home, or visit your nearest Ayushman Arogya Mandir. Medicines are also free every month.",
     "2026-08-05", "General", "Normal", "", "benefits.html", "No", "Yes", "Active", 3, ""),

    ("NOT04", "102 and 108 ambulances are free — including the ride home",
     "Dial 102 for pregnant women and newborns, and 108 for any medical emergency or accident. The ambulance comes to you, the service is free, and it drops you back home after discharge. Nobody may charge you for it.",
     "2026-07-28", "General", "Normal", "", "benefits.html", "No", "No", "Active", 4, ""),

    ("NOT05", "TB treatment is free, and patients receive monthly nutrition support",
     "Testing and the complete course of TB medicines are free at every government facility, including for drug-resistant TB. Every notified patient also receives money in their bank account each month for nutrition throughout treatment.",
     "2026-07-20", "General", "Normal", "", "benefits.html", "No", "No", "Active", 5, ""),

    ("NOT06", "Quarterly District Health Society meeting",
     "The quarterly governing body meeting of the District Health Society, Sheikhpura will be chaired by the District Magistrate.",
     "2026-05-28", "Meeting", "Low", "", "", "No", "No", "Active", 6, "2026-09-30"),
]

# ───────────────────────── 11. CONTACT ─────────────────────────
CONTACT = [
    ("CON01", "Office of the Civil Surgeon, Sheikhpura", "Civil Surgeon", MARK,
     "Sadar Hospital Campus, Sheikhpura", "Sheikhpura", "Bihar", "811105", MARK, "", MARK,
     "10:30 AM – 5:00 PM (Mon–Sat, except public holidays)",
     "https://www.google.com/maps/search/Sadar+Hospital+Sheikhpura+Bihar", 1, "Active"),
    ("CON02", "District Programme Management Unit (DPMU)", "District Programme Manager", MARK,
     "Sadar Hospital Campus, Sheikhpura", "Sheikhpura", "Bihar", "811105", MARK, "", MARK,
     "10:30 AM – 5:00 PM (Mon–Sat)", "", 2, "Active"),
    ("CON03", "District Accounts Cell — NHM", "District Accounts Manager", MARK,
     "Sadar Hospital Campus, Sheikhpura", "Sheikhpura", "Bihar", "811105", MARK, "", MARK,
     "10:30 AM – 5:00 PM (Mon–Sat)", "", 3, "Active"),
    ("CON04", "District Monitoring & Evaluation Cell", "District M&E Officer", MARK,
     "Sadar Hospital Campus, Sheikhpura", "Sheikhpura", "Bihar", "811105", MARK, "", MARK,
     "10:30 AM – 5:00 PM (Mon–Sat)", "", 4, "Active"),
    ("CON05", "Health Helpline", "24x7 Toll Free", "—",
     "Statewide", "Sheikhpura", "Bihar", "811105", "104", "102 (Ambulance)", "",
     "24 hours", "", 5, "Active"),
]

# ───────────────────────── 12. FOOTER ─────────────────────────
FOOTER = [
    ("FT01", "about",   "About This Portal", "", "", "", 1, 1, "Active", "No"),
    ("FT02", "link",    "Quick Links", "Home",            "index.html",     "", 2, 1, "Active", "No"),
    ("FT03", "link",    "Quick Links", "Free Health Services", "benefits.html", "", 2, 2, "Active", "No"),
    ("FT16", "link",    "Quick Links", "Health Programmes", "programmes.html", "", 2, 6, "Active", "No"),
    ("FT04", "link",    "Quick Links", "Documents",       "documents.html", "", 2, 3, "Active", "No"),
    ("FT05", "link",    "Quick Links", "Notices",         "notices.html",   "", 2, 4, "Active", "No"),
    ("FT06", "link",    "Quick Links", "Contact Us",      "contact.html",   "", 2, 5, "Active", "No"),
    ("FT07", "link",    "Related Sites", "State Health Society, Bihar", "https://shs.bihar.gov.in/", "", 3, 1, "Active", "Yes"),
    ("FT08", "link",    "Related Sites", "National Health Mission",     "https://nhm.gov.in/", "", 3, 2, "Active", "Yes"),
    ("FT09", "link",    "Related Sites", "Sheikhpura District Portal",  "https://sheikhpura.nic.in/", "", 3, 3, "Active", "Yes"),
    ("FT10", "link",    "Related Sites", "BHAVYA Citizen Portal",       "https://mera.bhavyabiharhealth.in/", "", 3, 4, "Active", "Yes"),
    ("FT11", "contact", "Contact", "", "", "Office of the Civil Surgeon, Sadar Hospital Campus, Sheikhpura, Bihar - 811105", 4, 1, "Active", "No"),
    ("FT12", "legal",   "", "Disclaimer",     "disclaimer.html",     "", 5, 1, "Active", "No"),
    ("FT13", "legal",   "", "Privacy Policy", "privacy.html",        "", 5, 2, "Active", "No"),
    ("FT14", "legal",   "", "Accessibility Statement", "accessibility.html", "", 5, 3, "Active", "No"),
    ("FT15", "legal",   "", "Site Map",       "sitemap.html",        "", 5, 4, "Active", "No"),
]

# ───────────── 13. POST_CATEGORIES — taxonomy for News / Events / Updates ─────────────
# Colour is the chip colour on the public site. Admin may add rows freely.
PCATS = [
    ("PCAT01", "Health Campaign",        "health-campaign",        "#0F7B3E", "campaign",  1,  "Active"),
    ("PCAT02", "Awareness Program",      "awareness-program",      "#1B7FA8", "megaphone", 2,  "Active"),
    ("PCAT03", "Training",               "training",               "#6941C6", "training",  3,  "Active"),
    ("PCAT04", "Workshop",               "workshop",               "#6941C6", "workshop",  4,  "Active"),
    ("PCAT05", "Meeting",                "meeting",                "#475467", "meeting",   5,  "Active"),
    ("PCAT06", "Health Camp",            "health-camp",            "#0F7B3E", "camp",      6,  "Active"),
    ("PCAT07", "Vaccination Drive",      "vaccination-drive",      "#B54708", "vaccine",   7,  "Active"),
    ("PCAT08", "Maternal Health",        "maternal-health",        "#C11574", "maternal",  8,  "Active"),
    ("PCAT09", "Child Health",           "child-health",           "#C11574", "child",     9,  "Active"),
    ("PCAT10", "NCD",                    "ncd",                    "#B42318", "ncd",       10, "Active"),
    ("PCAT11", "Public Health",          "public-health",          "#1B7FA8", "public",    11, "Active"),
    ("PCAT12", "Digital Health",         "digital-health",         "#3538CD", "digital",   12, "Active"),
    ("PCAT13", "BHAVYA",                 "bhavya",                 "#3538CD", "bhavya",    13, "Active"),
    ("PCAT14", "ABDM / ABHA",            "abdm-abha",              "#3538CD", "abdm",      14, "Active"),
    ("PCAT15", "District Achievement",   "district-achievement",   "#0F7B3E", "award",     15, "Active"),
    ("PCAT16", "Government Initiative",  "government-initiative",  "#1B3A5C", "gov",       16, "Active"),
    ("PCAT17", "Important Announcement", "important-announcement", "#B42318", "alert",     17, "Active"),
    ("PCAT18", "Other",                  "other",                  "#475467", "dot",       18, "Active"),
]

# ───────────── 14. POSTS — What's New / Events / News ─────────────
# Status drives visibility: Published (live), Scheduled (live from Scheduled_Date),
# Draft and Archived (never public). Upcoming vs Past is DERIVED from Event_Start_Date,
# never stored, so no one has to remember to change it.
_LONG_1 = (
    "The District Health Society, Sheikhpura is observing World Breastfeeding Week with awareness "
    "activities across all six blocks. Activities include mothers' meetings at Ayushman Arogya Mandirs, "
    "counselling sessions on early initiation and exclusive breastfeeding, rallies by ASHA workers and "
    "Anganwadi sevikas, and orientation of frontline workers on IYCF practices.\n\n"
    "Block-level teams will report daily activity counts to the DPMU. Facility in-charges are requested "
    "to ensure that every delivery point displays IYCF messaging and that breastfeeding counselling is "
    "recorded in the mother's case sheet."
)
_LONG_2 = (
    "A district-level health camp will be organised at PHC Chewara covering general OPD, non-communicable "
    "disease screening, eye check-up, anaemia testing and ABHA registration.\n\n"
    "Specialist doctors from Sadar Hospital will be deputed for the day. Free medicines will be provided "
    "under the Free Drug Service Initiative. Community mobilisation will be carried out by ASHA workers "
    "in the preceding week."
)
_LONG_3 = (
    "Sheikhpura district has achieved full ABHA (Ayushman Bharat Health Account) generation coverage at all "
    "health sub-centres, with every HSC now onboarded to the BHAVYA platform and reporting daily.\n\n"
    "The district recorded consistent improvement in teleconsultation uptake and vitals capture over the "
    "review period. The Civil Surgeon has commended the block teams and Community Health Officers for "
    "sustained data entry discipline."
)
_LONG_4 = (
    "A two-day training programme on Comprehensive Primary Health Care service packages will be conducted "
    "for Community Health Officers posted at Ayushman Arogya Mandirs.\n\n"
    "Sessions will cover the twelve service packages, teleconsultation protocols, NCD screening workflow, "
    "and reporting through the BHAVYA platform. Participants must bring their facility login credentials."
)
_LONG_5 = (
    "The monthly review meeting of the District Health Society will be chaired by the District Magistrate. "
    "Progress of health services across all six blocks will be reviewed, including maternal and child health, "
    "immunisation, disease control programmes and the functioning of Ayushman Arogya Mandirs."
)
_LONG_6 = (
    "A special measles-rubella vaccination drive will be conducted across all blocks of Sheikhpura district "
    "targeting children in the 9 months to 5 years age group.\n\n"
    "Session sites will be organised at Anganwadi centres, schools and health facilities. Micro-plans have "
    "been shared with block teams. Cold chain readiness has been verified at all points."
)
_LONG_7 = (
    "Revised monthly reporting formats for Ayushman Arogya Mandirs take effect from this quarter. The updated "
    "format captures teleconsultation counts, NCD screening disaggregated by age band, and ABHA linkage status.\n\n"
    "The format is available in the Documents section of this portal. Block Programme Managers must ensure all "
    "facilities switch to the new format from the current reporting month."
)

POSTS = [
    # Post_ID, Slug, Title, Short_Description, Full_Description, Content_Type, Category_ID,
    # Featured_Image_URL, Event_Start_Date, Event_End_Date, Event_Time, Venue, Location,
    # External_URL, Attachment_URL, Attachment_Name, Published_Date, Scheduled_Date, Author,
    # Is_Featured, Status, Created_Date, Updated_Date
    ("POST001", "world-breastfeeding-week-2026", "World Breastfeeding Week 2026",
     "District-level awareness activities are being organised across Sheikhpura to promote breastfeeding and improve maternal and child health outcomes.",
     _LONG_1, "Event", "PCAT08", "", "2026-08-01", "2026-08-07", "10:00 AM",
     "All Blocks", "Sheikhpura District", "", "", "", "2026-07-25", "", "District Programme Manager",
     "Yes", "Published", "2026-07-25", "2026-07-25"),

    ("POST002", "district-health-camp-chewara-2026", "District Health Camp — PHC Chewara",
     "A full-day health camp offering general OPD, NCD screening, eye check-up, anaemia testing and ABHA registration.",
     _LONG_2, "Event", "PCAT06", "", "2026-08-25", "2026-08-25", "09:00 AM",
     "PHC Chewara", "Chewara Block, Sheikhpura", "", "", "", "2026-08-05", "", "District Programme Manager",
     "Yes", "Published", "2026-08-05", "2026-08-05"),

    ("POST003", "abha-full-coverage-all-hsc", "Sheikhpura Achieves Full ABHA Coverage Across All Health Sub-Centres",
     "Every health sub-centre in the district is now onboarded to the BHAVYA platform with daily reporting and full ABHA generation coverage.",
     _LONG_3, "News", "PCAT15", "", "", "", "",
     "", "Sheikhpura District", "https://mera.bhavyabiharhealth.in/", "", "", "2026-08-12", "", "District M&E Officer",
     "Yes", "Published", "2026-08-12", "2026-08-12"),

    ("POST004", "cho-cphc-training-august-2026", "Two-Day CHO Training on CPHC Service Packages",
     "Training for Community Health Officers on the twelve CPHC service packages, teleconsultation protocols and BHAVYA reporting.",
     _LONG_4, "Event", "PCAT03", "", "2026-09-10", "2026-09-11", "10:00 AM",
     "District Training Centre, Sadar Hospital Campus", "Sheikhpura", "", "", "", "2026-08-14", "", "District Programme Manager",
     "No", "Published", "2026-08-14", "2026-08-14"),

    ("POST005", "dhs-monthly-review-august-2026", "Monthly District Health Society Review Meeting",
     "Monthly review of health services across the district, chaired by the District Magistrate.",
     _LONG_5, "Event", "PCAT05", "", "2026-08-28", "2026-08-28", "11:00 AM",
     "Collectorate Conference Hall", "Sheikhpura", "", "", "", "2026-08-16", "", "District Programme Manager",
     "No", "Published", "2026-08-16", "2026-08-16"),

    ("POST006", "measles-rubella-vaccination-drive-2026", "Special Measles-Rubella Vaccination Drive",
     "District-wide MR vaccination drive for children aged 9 months to 5 years, with session sites at Anganwadi centres, schools and health facilities.",
     _LONG_6, "Event", "PCAT07", "", "2026-09-15", "2026-09-30", "09:00 AM",
     "All Session Sites", "Sheikhpura District", "", "", "", "2026-08-17", "", "District Immunisation Officer",
     "No", "Published", "2026-08-17", "2026-08-17"),

    ("POST007", "revised-aam-reporting-format-2026", "Revised Monthly Reporting Format for Ayushman Arogya Mandirs",
     "Updated AAM reporting format takes effect this quarter, capturing teleconsultation counts, disaggregated NCD screening and ABHA linkage status.",
     _LONG_7, "Update", "PCAT12", "", "", "", "",
     "", "Sheikhpura District", "", MARK, "AAM Monthly Report Format.xlsx", "2026-08-10", "", "District M&E Officer",
     "No", "Published", "2026-08-10", "2026-08-10"),

    # Scheduled — hidden from the public site until Scheduled_Date passes
    ("POST008", "nutrition-month-2026", "Rashtriya Poshan Maah 2026",
     "Month-long nutrition awareness campaign with community-based events, growth monitoring drives and IYCF counselling across all blocks.",
     "Rashtriya Poshan Maah will be observed throughout September with weekly themes covering growth monitoring, "
     "anaemia, complementary feeding and Poshan Vatikas. Block teams will conduct Village Health Sanitation and "
     "Nutrition Days with expanded coverage.",
     "Event", "PCAT02", "", "2026-09-01", "2026-09-30", "", "All Blocks", "Sheikhpura District",
     "", "", "", "", "2026-08-28 09:00", "District Programme Manager", "No", "Scheduled",
     "2026-08-17", "2026-08-17"),

    # Draft — never public
    ("POST009", "eye-donation-fortnight-2026", "Eye Donation Fortnight 2026 (draft)",
     "Draft announcement for the district observance of Eye Donation Fortnight.",
     "Draft content pending approval from the Civil Surgeon. Dates and venue to be confirmed.",
     "Event", "PCAT11", "", "", "", "", "", "Sheikhpura District",
     "", "", "", "", "", "District Programme Manager", "No", "Draft", "2026-08-17", "2026-08-17"),

    # Archived — past, kept for the record
    ("POST010", "world-tb-day-2026", "World TB Day 2026 Observed Across the District",
     "Awareness rallies, screening camps and Ni-kshay Mitra felicitation were held across all blocks to mark World TB Day.",
     "World TB Day was observed with awareness rallies, active case-finding camps and felicitation of Ni-kshay Mitras "
     "who have supported nutritional assistance to TB patients in the district.",
     "News", "PCAT01", "", "2026-03-24", "2026-03-24", "", "All Blocks", "Sheikhpura District",
     "", "", "", "2026-03-25", "", "District TB Officer", "No", "Archived", "2026-03-25", "2026-03-25"),
]

# ───────────── 15. POST_MEDIA — gallery images, one row per image ─────────────
# Deliberately relational: the alternative (a delimited list crammed into one cell)
# is exactly what the brief said not to do.
POST_MEDIA = [
    ("PM001", "POST001", "", "image", "Mothers' meeting at Ayushman Arogya Mandir", "", "", 1, "Active"),
    ("PM002", "POST001", "", "image", "ASHA workers' awareness rally",                "", "", 2, "Active"),
    ("PM003", "POST001", "", "image", "IYCF counselling session",                     "", "", 3, "Active"),
    ("PM004", "POST002", "", "image", "Health camp inauguration",                     "", "", 1, "Active"),
    ("PM005", "POST002", "", "image", "Doctor consultation",                          "", "", 2, "Active"),
    ("PM006", "POST002", "", "image", "Patient registration desk",                    "", "", 3, "Active"),
    ("PM007", "POST002", "", "image", "NCD screening counter",                        "", "", 4, "Active"),
    ("PM008", "POST003", "", "image", "CHO entering vitals on the BHAVYA platform",    "", "", 1, "Active"),
]


# ───────────────────────── CONTROLLED LISTS ─────────────────────────
LISTS = [
    ("Status",           "Active",                  "Row is published and visible on the website"),
    ("Status",           "Inactive",                "Row is hidden from the website but kept in the sheet"),
    ("Status",           "Archived",                "Row is shown only in archive views / older financial years"),
    ("Yes_No",           "Yes",                     ""),
    ("Yes_No",           "No",                      ""),
    ("Document_Type",    "Programme Guideline",     "Guideline for a single FMR head"),
    ("Document_Type",    "Format",                  "Reporting or data-entry format"),
    ("Document_Type",    "Report",                  ""),
    ("Document_Type",    "Circular",                ""),
    ("Document_Type",    "Letter",                  ""),
    ("Document_Type",    "Other",                   ""),
    ("File_Type",        "PDF",                     ""),
    ("File_Type",        "XLSX",                    ""),
    ("File_Type",        "DOCX",                    ""),
    ("File_Type",        "ZIP",                     ""),
    ("File_Type",        "PPTX",                    ""),
    ("File_Type",        "LINK",                    "External page rather than a file"),
    ("Priority",         "High",                    "Shown first, red flag"),
    ("Priority",         "Normal",                  ""),
    ("Priority",         "Low",                     ""),
    ("Notice_Category",  "Circular",                ""),
    ("Notice_Category",  "Meeting",                 ""),
    ("Notice_Category",  "Recruitment",             ""),
    ("Notice_Category",  "General",                 ""),
    ("Link_Type",        "Internal",                "Page inside this website"),
    ("Link_Type",        "External",                "Opens another website"),
    ("Link_Category",    "National",                ""),
    ("Link_Category",    "State",                   ""),
    ("Link_Category",    "District",                ""),
    ("Link_Category",    "Portal",                  ""),
    ("Benefit_Type",     "Cash Benefit",            "Money paid to the citizen"),
    ("Benefit_Type",     "Free Service",            "A service given without charge"),
    ("Benefit_Type",     "Free Medicine",           ""),
    ("Benefit_Type",     "Free Test",               "Laboratory or diagnostic test"),
    ("Benefit_Type",     "Free Transport",          "Ambulance or travel support"),
    ("Benefit_Type",     "Free Equipment",          "Spectacles, hearing aid, mobility aid"),
    ("Benefit_Type",     "Awareness / Counselling", ""),
    ("Content_Type",     "News",                    "Announcement, achievement or update — no event date"),
    ("Content_Type",     "Event",                   "Has a date and usually a venue"),
    ("Content_Type",     "Update",                  "Short operational update or instruction"),
    ("Post_Status",      "Draft",                   "Not visible on the website"),
    ("Post_Status",      "Published",               "Live on the website now"),
    ("Post_Status",      "Scheduled",               "Goes live automatically at Scheduled_Date"),
    ("Post_Status",      "Archived",                "Removed from listings, kept for the record"),
    ("Media_Type",       "image",                   "Gallery photo"),
    ("Media_Type",       "document",                "Attachment shown on the post detail page"),
    ("Section_Type",     "hero",                    "Top banner of the home page"),
    ("Section_Type",     "banner",                  "Announcement strip"),
    ("Section_Type",     "stat",                    "One number tile"),
    ("Section_Type",     "section",                 "Section heading + description"),
    ("Section_Type",     "richtext",                "Paragraph block"),
    ("Footer_Block",     "about",                   ""),
    ("Footer_Block",     "link",                    ""),
    ("Footer_Block",     "contact",                 ""),
    ("Footer_Block",     "legal",                   ""),
]

README = [
    ("1", "What this workbook is",
     "This is the complete content database for the Sheikhpura District Health PIP Portal. The website reads these sheets and renders itself. Change a value here and the website changes — no code editing."),
    ("2", "Golden rules",
     "One record per row. Never merge cells. Never delete or rename the header row (row 1). Never reorder or rename columns. Never reuse an ID."),
    ("3", "Turning a row off",
     "Set Status to Inactive. Do not delete rows — deleting breaks the ID references used by other sheets."),
    ("4", "Sheets you will edit most",
     "Notices (weekly), Documents and PIP_Documents (whenever a file is published), Programs_FMR (once a year), Financial_Years (once a year)."),
    ("5", "Adding a new financial year",
     "Step 1: add one row to Financial_Years and set Is_Current=Yes on it and No on the old one. Step 2: copy last year's rows in Programs_FMR, change Year_ID to the new year, then add/remove FMR codes as per the new RoP. Step 3: add document rows. Nothing else."),
    ("6", "Why FMR codes repeat per year",
     "Because they genuinely change. FY 2025-26 had 49 FMR heads; FY 2026-27 has 59. Storing them once would silently show the wrong list for older years."),
    ("7", "Where files live",
     "Upload PDFs/XLSX to the district Google Drive folder, set sharing to 'Anyone with the link — Viewer', then paste the link into File_URL. Any public https:// link works."),
    ("8", "NEEDS MANUAL INPUT",
     "Cells with this red text are placeholders. The website hides or disables anything still marked this way. Replace them with real values before go-live."),
    ("9", "Validation",
     "The website validates every row on load. Rows with a missing required field are skipped and listed on the Admin dashboard — the site never breaks because of a bad row."),
    ("10", "Publishing changes",
     "Edits appear on the website within the cache window (default 30 minutes), or immediately after clicking 'Clear cache & sync now' on the Admin dashboard."),
]


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet(wb, "README", ["S.No", "Topic", "Instruction"], README,
          {"S.No": 8, "Topic": 32, "Instruction": 118})

    sheet(wb, "Settings",
          ["Setting_Key", "Setting_Label", "Setting_Value", "Group", "Notes"], SETTINGS,
          {"Setting_Key": 26, "Setting_Label": 26, "Setting_Value": 62, "Group": 14, "Notes": 46})

    sheet(wb, "Navigation",
          ["Menu_ID", "Menu_Label_EN", "Menu_Label_HI", "URL", "Parent_Menu_ID",
           "Display_Order", "Is_Active", "Link_Type", "Icon", "Target"], NAV,
          {"URL": 34, "Menu_Label_EN": 20, "Menu_Label_HI": 20})

    sheet(wb, "Financial_Years",
          ["Year_ID", "Financial_Year", "Display_Name", "Start_Year", "End_Year",
           "Is_Current", "Status", "Display_Order"], FY)

    sheet(wb, "Program_Categories",
          ["Category_ID", "Category_Name", "Short_Name", "Description", "Icon",
           "Display_Order", "Status"], CATS,
          {"Category_Name": 46, "Description": 78})

    sheet(wb, "Programs_FMR",
          ["Program_ID", "Year_ID", "Category_ID", "FMR_Code", "Program_Name",
           "Program_Name_HI", "Program_Description", "Nodal_Officer",
           "Display_Order", "Status"], PROGRAMS,
          {"Program_Name": 46, "Program_Description": 74,
           "Nodal_Officer": 30, "Program_Name_HI": 18})

    sheet(wb, "Documents",
          ["Document_ID", "Year_ID", "Category_ID", "Program_ID", "Document_Title",
           "Document_Type", "Description", "File_URL", "File_Type", "File_Size_MB",
           "Upload_Date", "Display_Order", "Status", "Is_Featured"], DOCUMENTS,
          {"Document_Title": 46, "Description": 66, "File_URL": 30, "Document_Type": 22})

    sheet(wb, "Home_Content",
          ["Section_Key", "Section_Type", "Title_EN", "Title_HI", "Subtitle",
           "Body_Text", "Icon", "Link_URL", "Link_Label", "Display_Order", "Status"], HOME,
          {"Body_Text": 96, "Title_EN": 30, "Subtitle": 40, "Section_Key": 18})

    sheet(wb, "Important_Links",
          ["Link_ID", "Link_Name", "URL", "Description", "Icon", "Category",
           "Display_Order", "Status", "Is_External", "Show_In_Footer", "Show_On_Home"], LINKS,
          {"Link_Name": 34, "URL": 42, "Description": 46})

    sheet(wb, "Notices",
          ["Notice_ID", "Title", "Description", "Notice_Date", "Category", "Priority",
           "Attachment_URL", "External_URL", "Is_Featured", "Is_New", "Status",
           "Display_Order", "Expiry_Date"], NOTICES,
          {"Title": 46, "Description": 96, "External_URL": 26})

    sheet(wb, "Contact_Information",
          ["Contact_ID", "Office_Name", "Designation", "Person_Name", "Address",
           "District", "State", "PIN", "Phone", "Alt_Phone", "Email", "Office_Hours",
           "Google_Maps_URL", "Display_Order", "Status"], CONTACT,
          {"Office_Name": 42, "Address": 38, "Office_Hours": 44, "Google_Maps_URL": 34})

    sheet(wb, "Footer",
          ["Footer_ID", "Block_Type", "Block_Title", "Label", "URL", "Content_Text",
           "Column_Number", "Display_Order", "Status", "Is_External"], FOOTER,
          {"Label": 34, "URL": 40, "Content_Text": 66, "Block_Title": 20})

    sheet(wb, "Post_Categories",
          ["Category_ID", "Category_Name", "Slug", "Colour", "Icon",
           "Display_Order", "Status"], PCATS,
          {"Category_Name": 26, "Slug": 24})

    sheet(wb, "Posts",
          ["Post_ID", "Slug", "Title", "Short_Description", "Full_Description",
           "Content_Type", "Category_ID", "Featured_Image_URL", "Event_Start_Date",
           "Event_End_Date", "Event_Time", "Venue", "Location", "External_URL",
           "Attachment_URL", "Attachment_Name", "Published_Date", "Scheduled_Date",
           "Author", "Is_Featured", "Status", "Created_Date", "Updated_Date"], POSTS,
          {"Slug": 36, "Title": 52, "Short_Description": 74, "Full_Description": 100,
           "Venue": 34, "Location": 24, "Author": 28, "Featured_Image_URL": 30,
           "External_URL": 30, "Attachment_URL": 24, "Attachment_Name": 28})

    sheet(wb, "Post_Media",
          ["Media_ID", "Post_ID", "Media_URL", "Media_Type", "Caption",
           "File_Name", "File_Size_KB", "Display_Order", "Status"], POST_MEDIA,
          {"Media_URL": 34, "Caption": 44, "File_Name": 26})

    sheet(wb, "Program_Benefits", BENEFIT_COLUMNS, BENEFITS,
          {"Benefit_Title": 52, "Benefit_Title_HI": 46, "Benefit_Description": 96,
           "Benefit_Description_HI": 96, "Who_Is_Eligible": 56, "Where_To_Avail": 56,
           "Documents_Required": 44, "Amount": 42, "Benefit_Type": 22})

    ws = sheet(wb, "_Lists", ["List_Name", "Value", "Meaning"], LISTS,
               {"List_Name": 20, "Value": 30, "Meaning": 60})
    ws.sheet_properties.tabColor = "98A2B3"

    # data-validation dropdowns driven by _Lists (survive the Google Sheets import)
    def add_dv(sheet_name, col_letter, list_name, count):
        vals = [v for n, v, _ in LISTS if n == list_name]
        first = next(i for i, (n, _, _) in enumerate(LISTS, start=2) if n == list_name)
        last = first + len(vals) - 1
        dv = DataValidation(type="list", formula1=f"=_Lists!$B${first}:$B${last}",
                            allow_blank=True, showDropDown=False)
        wb[sheet_name].add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{count + 1}")

    add_dv("Programs_FMR",       "J", "Status",          len(PROGRAMS))
    add_dv("Documents",          "M", "Status",          len(DOCUMENTS))
    add_dv("Documents",          "F", "Document_Type",   len(DOCUMENTS))
    add_dv("Documents",          "I", "File_Type",       len(DOCUMENTS))
    add_dv("Notices",            "K", "Status",          len(NOTICES))
    add_dv("Notices",            "F", "Priority",        len(NOTICES))
    add_dv("Notices",            "E", "Notice_Category", len(NOTICES))
    add_dv("Important_Links",    "H", "Status",          len(LINKS))
    add_dv("Important_Links",    "F", "Link_Category",   len(LINKS))
    add_dv("Navigation",         "H", "Link_Type",       len(NAV))
    add_dv("Financial_Years",    "G", "Status",          len(FY))
    add_dv("Program_Categories", "G", "Status",          len(CATS))
    add_dv("Contact_Information","O", "Status",          len(CONTACT))
    add_dv("Footer",             "I", "Status",          len(FOOTER))
    add_dv("Home_Content",       "K", "Status",          len(HOME))
    add_dv("Post_Categories",    "G", "Status",          len(PCATS))
    add_dv("Posts",              "F", "Content_Type",    len(POSTS))
    add_dv("Posts",              "U", "Post_Status",     len(POSTS))
    add_dv("Posts",              "T", "Yes_No",          len(POSTS))
    add_dv("Post_Media",         "D", "Media_Type",      len(POST_MEDIA))
    add_dv("Post_Media",         "I", "Status",          len(POST_MEDIA))
    add_dv("Program_Benefits",   "H", "Benefit_Type",    len(BENEFITS))
    add_dv("Program_Benefits",   "O", "Status",          len(BENEFITS))

    wb.save(OUT)

    # ── integrity checks ──
    print(f"Saved: {OUT}\n")
    fy_ids = {r[0] for r in FY}
    cat_ids = {r[0] for r in CATS}
    prog_ids = {r[0] for r in PROGRAMS}
    errs = []

    def uniq(name, ids):
        if len(ids) != len(set(ids)):
            errs.append(f"{name}: duplicate IDs")

    uniq("Financial_Years", [r[0] for r in FY])
    uniq("Program_Categories", [r[0] for r in CATS])
    uniq("Programs_FMR", [r[0] for r in PROGRAMS])
    uniq("Documents", [r[0] for r in DOCUMENTS])
    uniq("Notices", [r[0] for r in NOTICES])
    uniq("Important_Links", [r[0] for r in LINKS])
    uniq("Post_Categories", [r[0] for r in PCATS])
    uniq("Posts", [r[0] for r in POSTS])
    uniq("Post_Media", [r[0] for r in POST_MEDIA])

    # Posts: slug must be unique and URL-safe; FKs and status must be valid
    pcat_ids = {r[0] for r in PCATS}
    post_ids = {r[0] for r in POSTS}
    slugs = [r[1] for r in POSTS]
    if len(slugs) != len(set(slugs)):
        errs.append("Posts: duplicate Slug — slugs are the public URL and must be unique")
    for r in POSTS:
        pid, slug, ctype, cat, status = r[0], r[1], r[5], r[6], r[20]
        if not re.fullmatch(r"[a-z0-9]+(?:-[a-z0-9]+)*", slug):
            errs.append(f"Posts {pid}: slug '{slug}' is not URL-safe (lowercase, digits, hyphens only)")
        if cat not in pcat_ids:
            errs.append(f"Posts {pid}: bad Category_ID {cat}")
        if ctype not in {"News", "Event", "Update"}:
            errs.append(f"Posts {pid}: bad Content_Type {ctype}")
        if status not in {"Draft", "Published", "Scheduled", "Archived"}:
            errs.append(f"Posts {pid}: bad Status {status}")
        if status == "Scheduled" and not str(r[17]).strip():
            errs.append(f"Posts {pid}: Status=Scheduled requires a Scheduled_Date")
        if status == "Published" and not str(r[16]).strip():
            errs.append(f"Posts {pid}: Status=Published requires a Published_Date")
        if r[8] and r[9] and str(r[9]) < str(r[8]):
            errs.append(f"Posts {pid}: Event_End_Date is before Event_Start_Date")
    uniq("Program_Benefits", [r[0] for r in BENEFITS])
    all_codes = {r[3] for r in PROGRAMS}
    for r in BENEFITS:
        if r[1] not in all_codes:
            errs.append(f"Program_Benefits {r[0]}: FMR_Code {r[1]} does not exist in Programs_FMR")
        if r[2] and r[2] not in fy_ids:
            errs.append(f"Program_Benefits {r[0]}: bad Year_ID {r[2]}")
        if r[7] not in set(BENEFIT_TYPES):
            errs.append(f"Program_Benefits {r[0]}: bad Benefit_Type {r[7]}")
        if not str(r[3]).strip() or not str(r[5]).strip():
            errs.append(f"Program_Benefits {r[0]}: missing title or description")

    for r in POST_MEDIA:
        if r[1] not in post_ids:
            errs.append(f"Post_Media {r[0]}: bad Post_ID {r[1]}")

    for r in PROGRAMS:
        if r[1] not in fy_ids: errs.append(f"Programs_FMR {r[0]}: bad Year_ID {r[1]}")
        if r[2] not in cat_ids: errs.append(f"Programs_FMR {r[0]}: bad Category_ID {r[2]}")
    for r in DOCUMENTS:
        if r[1] and r[1] not in fy_ids: errs.append(f"Documents {r[0]}: bad Year_ID")
        if r[2] and r[2] not in cat_ids: errs.append(f"Documents {r[0]}: bad Category_ID")
        if r[3] and r[3] not in prog_ids: errs.append(f"Documents {r[0]}: bad Program_ID {r[3]}")

    # one FMR code may not repeat within the same year
    seen = set()
    for r in PROGRAMS:
        k = (r[1], r[3])
        if k in seen: errs.append(f"Duplicate FMR {r[3]} in {r[1]}")
        seen.add(k)

    cur = [r for r in FY if r[5] == "Yes"]
    if len(cur) != 1: errs.append("Financial_Years: exactly one row must have Is_Current=Yes")

    counts = {}
    for r in PROGRAMS:
        counts[r[1]] = counts.get(r[1], 0) + 1
    print("Row counts")
    print(f"  Settings            {len(SETTINGS):>4}")
    print(f"  Navigation          {len(NAV):>4}")
    print(f"  Financial_Years     {len(FY):>4}")
    print(f"  Program_Categories  {len(CATS):>4}")
    print(f"  Programs_FMR        {len(PROGRAMS):>4}   " +
          "  ".join(f"{k}={v}" for k, v in sorted(counts.items(), reverse=True)))
    print(f"  Documents           {len(DOCUMENTS):>4}")
    print(f"  Home_Content        {len(HOME):>4}")
    print(f"  Important_Links     {len(LINKS):>4}")
    print(f"  Notices             {len(NOTICES):>4}")
    print(f"  Contact_Information {len(CONTACT):>4}")
    print(f"  Footer              {len(FOOTER):>4}")
    pstat = {}
    for r in POSTS:
        pstat[r[20]] = pstat.get(r[20], 0) + 1
    print(f"  Post_Categories     {len(PCATS):>4}")
    print(f"  Posts               {len(POSTS):>4}   " +
          "  ".join(f"{k}={v}" for k, v in sorted(pstat.items())))
    print(f"  Post_Media          {len(POST_MEDIA):>4}")
    bcodes = len({r[1] for r in BENEFITS})
    print(f"  Program_Benefits    {len(BENEFITS):>4}   across {bcodes} FMR codes")
    print(f"  _Lists              {len(LISTS):>4}")
    print(f"\nIntegrity: {'PASS — no errors' if not errs else 'FAIL'}")
    for e in errs:
        print("   !", e)


if __name__ == "__main__":
    main()
